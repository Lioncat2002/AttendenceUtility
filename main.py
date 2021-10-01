import openpyxl

fullstudents=openpyxl.load_workbook("test_fullStudents.xlsx")

fullsheet=fullstudents.active
perdaystudents=openpyxl.load_workbook("test_student_perday.xlsx")
perdaysheet=perdaystudents.active
allstudents=[fullsheet[f"B{i}"].value for i in range(2,fullsheet.max_row+1)]
presentstudents=[perdaysheet[f"A{i}"].value for i in range(2,perdaysheet.max_row+1)]
absentees=[]

for i in allstudents:
    if i not in presentstudents:
        
        for j in range(2,fullsheet.max_row+1):
            if fullsheet[f"B{j}"].value==i:
                absentees.append([fullsheet[f"A{j}"].value,i])
                break
print(absentees)
