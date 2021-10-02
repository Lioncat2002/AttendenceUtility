import openpyxl
import sys
fullstudents=openpyxl.load_workbook(sys.argv[1])

fullsheet=fullstudents.active
perdaystudents=openpyxl.load_workbook(sys.argv[2])
perdaysheet=perdaystudents.active
allstudents=[fullsheet[f"A{i}"].value for i in range(1,fullsheet.max_row+1)]

presentstudents=[perdaysheet[f"H{i}"].value for i in range(1,perdaysheet.max_row+1)]
absentees=[]

for i in allstudents:
    if i not in presentstudents:
        
        for j in range(1,fullsheet.max_row+1):
            if fullsheet[f"A{j}"].value==i:
                absentees.append([fullsheet[f"B{j}"].value,i])
                break
print(absentees)
